from openpyxl import load_workbook
import os

file_path = r"D:\Search\dashboard_python\03 Automatizando Tarefas com Python\0. Onboarding\Excel com Python\exemplo.xlsx"

if not os.path.exists(file_path):
    raise FileNotFoundError(f"Arquivo não encontrado: {file_path}")

wb = load_workbook(file_path)

print(wb.sheetnames)
sheet = wb["Sheet1"]

formula = "=SUM(C2:C7)"
sheet['C8'] = formula
wb.save(file_path)


from openpyxl.formula.translate import Translator
sheet['D8'] = Translator(formula, origin="C8").translate_formula("D8")
sheet['D8'].value

wb.save(file_path)


# Você pode consultar as fórmulas disponíveis aqui
from openpyxl.utils import FORMULAE
FORMULAE
